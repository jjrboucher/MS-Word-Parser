from openpyxl import load_workbook
from openpyxl import Workbook
import os


def write_worksheet(excel_filepath, worksheet_name, headers, rows_of_data):
    """
    This function will write data to a worksheet within an Excel file.
    It will iterate over the list rows_of_data passed ot it.
    In some cases, there will only be one row. In others, there will be several.
    In either case, the function works.

    param: excel_filepath
    param: docx_file
    param: worksheet_name
    param: headers
    param: rows_of_data - Must be a two-dimensional list, as Excel requires a list to write a row to a worksheet.

    return: True/False depending on if it was successful in writing the worksheet.
    """
    try:
        if os.path.exists(excel_filepath):  # if the file exists, open it.
            workbook = load_workbook(excel_filepath)
        else:  # otherwise, create it
            workbook = Workbook()

        if worksheet_name in workbook.sheetnames:  # if the worksheet metadata already exists, select it.
            worksheet = workbook[worksheet_name]
        else:
            # Create the worksheet
            worksheet = workbook.create_sheet(title=worksheet_name)
            worksheet.append(headers)  # Writes the headings to the spreadsheet

        for row in rows_of_data:  # write rows to the worksheet.
            worksheet.append(row)  # write the row

    except Exception as function_error:
        print(f"An error occurred while writing to Excel: {function_error}")
        return False  # Lets the main script calling this function know that it experienced an error writing to Excel.

    # Remove the default sheet created by openpyxl
    default_sheet = workbook.active
    if default_sheet.title == "Sheet":
        workbook.remove(default_sheet)

    workbook.save(excel_filepath)  # save the file

    return True  # Lets the main script know that it was successful in writing to Excel.
