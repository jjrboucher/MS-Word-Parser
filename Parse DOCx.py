####################################
# Written by Jacques Boucher
# jjrboucher@gmail.com
#
# Version Date: 16 September 2023
#
# Written in Python 3.11
#
# ********** Description **********
#
# Script will open a windows dialog to allow you to select a DOCx file.
# The script does not attempt to validate the file.
# A docx file is nothing more than a ZIP file, hence why this script uses the zipfile library.
#
# It will extract the results to a file called docx-artifacts.xlsx as defined by the variable excel_file_path at the
# start of the main part of the script.
# If the file does not exist, it creates it. If the file does exist, it appends to it.
# The file will be located in the folder where the script is executed from.
# If executing from the GUI by double-clicking on the .py file, it should be stored in that same folder.
# If executing it from the command line, it will create it in whichever folder you are in when executing it.
#
# This allows you to run this repeatedly against many DOCx file for an investigation and compare them.
# You can then copy/move/rename the default file, docx-artifacts.xlsx to another file name for your case.
#
# Processes that this script will do:
#
# 1 - It will extract a list of all the files within the zip file and save it to a worksheet called XML_files.
#     In this worksheet, it will save the following information to a row:
#     "File Name", "XML", "Size (bytes)", "MD5Hash"
#
# 2 - It will extract all the unique RSIDs from the file word/settings.xml and write it to a worksheet
#     called doc_summary.
#     In this worksheet, it will save the following information to a row:
#     "File Name", "Unique RSIDs", "RSID Root", "<w:p> tags", "<w:r> tags", "<w:t> tags"
#     Where "Unique RSID" is a numerical count of the # of RSIDs in the file.
#
#     What is an RSID (Revision Save ID)?
#     See https://learn.microsoft.com/en-us/dotnet/api/documentformat.openxml.wordprocessing.rsid?view=openxml-2.8.1
#
# 3 - It will extract all the unique RSIDs from the file word/settings.xml and write it to a worksheet called rsids.
#     In this worksheet, it will save the following information to rows (one for each unique RSID):
#     "File Name", "RSID"
#
# 3 - It will extract all known relevant metadata from the files docProps/app.xml and docProps/core.xml
#     and write it to a worksheet called metadata.
#     In this worksheet, it will save the following information to a row:
#     "File Name", "Author", "Created Date","Last Modified By","Modified Date","Last Printed Date","Manager","Company",
#     "Revision","Total Editing Time","Pages","Paragraphs","Lines","Words","Characters","Characters With Spaces",
#     "Title","Subject","Keywords","Description","Application","App Version","Template","Doc Security","Category",
#     "contentStatus"
#
#
# ********** Dependencies **********
#
# If running the script on a Linux system, you may need to install python-tk. You can do this with the following
# command on a Debian (e.g. Ubuntu) system from the terminal window:
# sudo apt-get install python3-tk
#
# Whether running on Linux, Mac, or Windows, you may need to install some of the libraries if they are not included in
# your installation of Python 3.
# In particular, you may need to install openpyxl and hashlib. You can do so as follows from a terminal window:
#
# pip3 install openpyxl
# pip3 install hashlib
#
# If any other libraries are missing when trying to execute the script, install those in the same manner.
#
###################################

import os
import zipfile
import tkinter as tk
import hashlib
from tkinter import filedialog
import re
from openpyxl import load_workbook
from openpyxl import Workbook


def extract_rsids_from_xml(xmlcontent):
    try:
        all_rsids = []
        pattern = r'<w:rsid(?:[^>]*)/>'
        matches = re.findall(pattern, xmlcontent)  # Find all RSIDs

        print("Processing word/settings.xml for RSIDs")
        for match in matches:
            rsid_match = re.search(r'<w:rsid w:val="([^"]*)"', match)  # Loops through them
            if rsid_match:
                all_rsids.append(rsid_match.group(1))  # Appends it to the list

        print("Processing word/settings.xml for rsidRoot.")
        rsid_root = re.search(r'<w:rsidRoot w:val="([^"]*)"', xmlcontent)

        if rsid_root is None:
            rsid_root = ""
        else:
            rsid_root = rsid_root.group(1)

        return all_rsids, rsid_root

    except Exception as function_error:
        print(f"An error occurred while extracting RSIDs: {function_error}")
        return []  # if it can't find any RSID (that should never happen), it returns an empty list.


def extract_from_app_xml(xmlcontent):
    # extract relevant metadata from app.xml file using a GREP expression
    app_xml = {"template": re.search(r'<Template>(.*?)</Template>', xmlcontent),
               "totalTime": re.search(r'<TotalTime>(.*?)</TotalTime>', xmlcontent),
               "pages": re.search(r'<Pages>(.*?)</Pages>', xmlcontent),
               "words": re.search(r'<Words>(.*?)</Words>', xmlcontent),
               "characters": re.search(r'<Characters>(.*?)</Characters>', xmlcontent),
               "application": re.search(r'<Application>(.*?)</Application>', xmlcontent),
               "docSecurity": re.search(r'<DocSecurity>(.*?)</DocSecurity>', xmlcontent),
               "lines": re.search(r'<Lines>(.*?)</Lines>', xmlcontent),
               "paragraphs": re.search(r'<Paragraphs>(.*?)</Paragraphs>', xmlcontent),
               "charactersWithSpaces": re.search(r'<CharactersWithSpaces>(.*?)</CharactersWithSpaces>', xmlcontent),
               "appVersion": re.search(r'<AppVersion>(.*?)</AppVersion>', xmlcontent),
               "manager": re.search(r'<Manager>(.*?)</Manager>', xmlcontent),
               "company": re.search(r'<Company>(.*?)</Company>', xmlcontent)}

    print("Processing docProps/app.xml for metadata.")
    for key, value in app_xml.items():  # check the results of the GREP searches
        if value is None:  # if no hit, assign empty value
            app_xml[key] = ""
        else:  # if a hit, extract group(1) from the search hit
            app_xml[key] = app_xml[key].group(1)

    return app_xml


def extract_from_core_xml(xmlcontent):
    # extract relevant metadata from core.xml file using a GREP expression
    core_xml = {"title": re.search(r'<dc:title>(.*?)</dc:title>', xmlcontent),
                "subject": re.search(r'<dc:subject>(.*?)</dc:subject>', xmlcontent),
                "creator": re.search(r'<dc:creator>(.*?)</dc:creator>', xmlcontent),
                "keywords": re.search(r'<cp:keywords>(.*?)</cp:keywords>', xmlcontent),
                "description": re.search(r'<dc:description>(.*?)</dc:description>', xmlcontent),
                "revision": re.search(r'<cp:revision>(.*?)</cp:revision>', xmlcontent),
                "created": re.search(r'<dcterms:created.*?>(.*?)</dcterms:created>', xmlcontent),
                "modified": re.search(r'<dcterms:modified.*?>(.*?)</dcterms:modified>', xmlcontent),
                "lastModifiedBy": re.search(r'<cp:lastModifiedBy>(.*?)</cp:lastModifiedBy>', xmlcontent),
                "lastPrinted": re.search(r'<cp:lastPrinted>(.*?)</cp:lastPrinted>', xmlcontent),
                "category": re.search(r'<cp:category>(.*?)</cp:category>', xmlcontent),
                "contentStatus": re.search(r'<cp:contentStatus>(.*?)</cp:contentStatus>', xmlcontent)}

    print("Processing docProps/core.xml for metadata.")
    for key, value in core_xml.items():  # check the results of the GREP searches
        if value is None:  # if no hit, assign empty value
            core_xml[key] = ""
        else:  # if a hit, extract group(1) from the search hit
            core_xml[key] = core_xml[key].group(1)
    return core_xml


def list_of_xml_files(filename_path, file_name):
    print("Processing word/document.xml for list of XML files.")
    with zipfile.ZipFile(filename_path, 'r') as zip_file:
        # list content of the DOCx file
        xml_files = []
        for file_info in zip_file.infolist():
            with zipfile.ZipFile(filename_path, 'r') as zip_ref:
                with zip_ref.open(file_info.filename) as xml_file:
                    md5hash = hashlib.md5(xml_file.read()).hexdigest()
            xml_files.append([file_name, file_info.filename, file_info.file_size, md5hash])
        return xml_files


def extract_tags_from_document_xml(xmlcontent):
    # extract relevant artifacts from document.xml
    print("Processing word/document.xml to count # of <w:p>, <w:r>, and <w:t> tags.")
    document_xml = {"paragraphs": len(re.findall(r'</w:p>', xmlcontent)),
                    "runs": len(re.findall(r'</w:r>', xmlcontent)),
                    "text": len(re.findall(r'</w:t>', xmlcontent))}
    return document_xml


def write_to_excel(excel_filepath, file_name, xml_files, all_rsids, document_summary, rsid_root,
                   all_metadata):
    try:
        if os.path.exists(excel_filepath):  # if the file exists, open it.
            workbook = load_workbook(excel_filepath)
        else:  # otherwise, create it
            workbook = Workbook()

        # List of files in DOCx document
        if "XML_files" in workbook.sheetnames:  # if the worksheet XML_files already exists, select it.
            worksheet = workbook["XML_files"]
        else:
            # Create the worksheet "XML_files"
            worksheet = workbook.create_sheet(title="XML_files")
            worksheet.append(["File Name", "XML", "Size (bytes)", "MD5Hash"])

        for msword_file, xml_file, file_size, md5hash in xml_files:  # Loop through all the embedded files
            # Write a row to the spreadsheet for each embedded file.
            worksheet.append([msword_file, xml_file, file_size, md5hash])

        print(f"List of XML files along with size and hash appended to worksheet 'XML_files'")

        # Summary worksheet of # of RSIDs in a document
        if "doc_summary" in workbook.sheetnames:  # if the worksheet doc_summary already exits, select it.
            worksheet = workbook["doc_summary"]
        else:
            # Create the worksheet "doc_summary"
            worksheet = workbook.create_sheet(title="doc_summary")
            worksheet.append(["File Name", "Unique RSIDs", "RSID Root", "<w:p> tags", "<w:r> tags", "<w:t> tags"])

        worksheet.append([file_name, len(rsids), rsid_root, document_summary["paragraphs"],
                          document_summary["runs"], document_summary["text"]])

        print(f"Document summary appended to worksheet 'doc_summary'")

        # Check if the worksheet "rsids" already exists
        if "rsids" in workbook.sheetnames:  # if the worksheet rsids already exists, select it.
            worksheet = workbook["rsids"]
        else:
            # Create the worksheet "rsids"
            worksheet = workbook.create_sheet(title="rsids")
            worksheet.append(["File Name", "RSID"])

        for rsid in set(all_rsids):
            worksheet.append([file_name, rsid])

        print(f"Unique RSIDs appended to worksheet 'rsids'")

        # Check if the worksheet "metadata" already exists
        if "metadata" in workbook.sheetnames:  # if the worksheet metadata already exists, select it.
            worksheet = workbook["metadata"]
        else:
            # Create the worksheet "metadata"
            worksheet = workbook.create_sheet(title="metadata")
            headings = list(all_metadata.keys())  # Adds the keys as column headings to a list
            headings.insert(0, "File Name")  # Adds column heading "File Name" at the start of the list
            worksheet.append(headings)  # Writes the headings to the spreadsheet

        metadata = list(all_metadata.values())  # Adds values to the list
        metadata.insert(0, file_name)  # Adds the file name to the start of the list
        worksheet.append(metadata)  # Writes the metadata to the spreadsheet

        print(f"Metadata appended to worksheet 'metadata'")

        # Remove the default sheet created by openpyxl
        default_sheet = workbook.active
        if default_sheet.title == "Sheet":
            workbook.remove(default_sheet)

        workbook.save(excel_file_path)  # save the file

        print(f"Results written to {excel_file_path}.")

    except Exception as function_error:
        print(f"An error occurred while writing to Excel: {function_error}")


if __name__ == "__main__":

    # Output file - same path as where the script is run. It will create it if it does not exist,
    # or append to it if it does.
    excel_file_path = "docx-artifacts.xlsx"

    root = tk.Tk()
    root.withdraw()  # Hide the main window

    zip_file_path = filedialog.askopenfilename(title="Select DOCx File", filetypes=[("DOCx Files", "*.docx")])
    if not zip_file_path:
        print("No DOCx file selected. Exiting.")
    else:

        filename = os.path.basename(zip_file_path)

        # list of XML file in DOCx
        XMLFiles = list_of_xml_files(zip_file_path,
                                     filename)  # Executes the function to get a list of all XML files in DOCx file

        # parse word/settings.xml artifacts
        xml_file_path_within_zip = "word/settings.xml"  # Path of the XML file within the ZIP

        try:
            with zipfile.ZipFile(zip_file_path, 'r') as zipref:
                with zipref.open(xml_file_path_within_zip) as xmlFile:
                    xml_content = xmlFile.read().decode("utf-8")

                    rsids, rsidRoot = extract_rsids_from_xml(
                        xml_content)  # Executes the function to get all unique RSIDs.

        except FileNotFoundError:
            print(f"File '{xml_file_path_within_zip}' not found in the ZIP archive.")
        except Exception as e:
            print(f"An error occurred: {e}")

        # parse docProps/app.xml artifacts
        xml_file_path_within_zip = "docProps/app.xml"  # Path of the XML file within the ZIP

        try:
            with zipfile.ZipFile(zip_file_path, 'r') as zipref:
                with zipref.open(xml_file_path_within_zip) as xmlFile:
                    xml_content = xmlFile.read().decode("utf-8")
                    app_xml_metadata = extract_from_app_xml(
                        xml_content)  # Executes the function to get metadata from app.xml

        except FileNotFoundError:
            print(f"File '{xml_file_path_within_zip}' not found in the ZIP archive.")
        except Exception as e:
            print(f"An error occurred: {e}")

        # parse docProps/core.xml artifacts
        xml_file_path_within_zip = "docProps/core.xml"  # Path of the XML file within the ZIP
        try:
            with zipfile.ZipFile(zip_file_path, 'r') as zipref:
                with zipref.open(xml_file_path_within_zip) as xmlFile:
                    xml_content = xmlFile.read().decode("utf-8")
                    core_xml_metadata = extract_from_core_xml(
                        xml_content)  # Executes the function to get the metadata from core.xml

        except FileNotFoundError:
            print(f"File '{xml_file_path_within_zip}' not found in the ZIP archive.")
        except Exception as e:
            print(f"An error occurred: {e}")

        # parse word/document.xml artifacts
        xml_file_path_within_zip = "word/document.xml"  # Path of the XML file within the ZIP
        try:
            with zipfile.ZipFile(zip_file_path, 'r') as zipref:
                with zipref.open(xml_file_path_within_zip) as xmlFile:
                    xml_content = xmlFile.read().decode("utf-8")
                    documentXMLTagSummary = extract_tags_from_document_xml(xml_content)
                    # Executes the function to get the metadata from core.xml

        except FileNotFoundError:
            print(f"File '{xml_file_path_within_zip}' not found in the ZIP archive.")
        except Exception as e:
            print(f"An error occurred: {e}")

        # The keys will be used as the column heading in the spreadsheet
        # The order they are in is the order that the columns will be in the spreadsheet
        # Corresponding values passed, resulting in a dictionary being passed called allMetadata
        # containing column headings and associated extracted metadata value.
        allMetadata = {"Author": core_xml_metadata["creator"],
                       "Created Date": core_xml_metadata["created"],
                       "Last Modified By": core_xml_metadata["lastModifiedBy"],
                       "Modified Date": core_xml_metadata["modified"],
                       "Last Printed Date": core_xml_metadata["lastPrinted"],
                       "Manager": app_xml_metadata["manager"],
                       "Company": app_xml_metadata["company"],
                       "Revision": core_xml_metadata["revision"],
                       "Total Editing Time": app_xml_metadata["totalTime"],
                       "Pages": app_xml_metadata["pages"],
                       "Paragraphs": app_xml_metadata["paragraphs"],
                       "Lines": app_xml_metadata["lines"],
                       "Words": app_xml_metadata["words"],
                       "Characters": app_xml_metadata["characters"],
                       "Characters With Spaces": app_xml_metadata["charactersWithSpaces"],
                       "Title": core_xml_metadata["title"],
                       "Subject": core_xml_metadata["subject"],
                       "Keywords": core_xml_metadata["keywords"],
                       "Description": core_xml_metadata["description"],
                       "Application": app_xml_metadata["application"],
                       "App Version": app_xml_metadata["appVersion"],
                       "Template": app_xml_metadata["template"],
                       "Doc Security": app_xml_metadata["docSecurity"],
                       "Category": core_xml_metadata["category"],
                       "Content Status": core_xml_metadata["contentStatus"]
                       }

        write_to_excel(excel_file_path, filename, XMLFiles, rsids, documentXMLTagSummary, rsidRoot,
                       allMetadata)
